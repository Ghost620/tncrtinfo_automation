#!/usr/bin/env python
# coding: utf-8

# In[5]:


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import time, os, random
import PySimpleGUI as sg
import random,time
from datetime import date
from datetime import datetime
from selenium.webdriver.firefox.options import Options
# options = Options()
# options.binary_location = r'C:\Program Files\Mozilla Firefox\firefox.exe'

today = str(date.today()).split('-')
theme_name_list = sg.theme_list()
court_names = {'mon':'montgomery','rut':'rutherford','rob':'robertson','sum':'sumner','will':'williamson','wil':'wilson'}
mode_names = {'gen' : 'General Sessions', 'cir':'Circuit Court'}

def courts_login():
    driver.find_element(By.LINK_TEXT, 'Subscription Login').click()
    driver.find_element(By.ID, 'cphContent_cphFormDetail_logmain_UserName').send_keys(username)
    driver.find_element(By.ID, 'cphContent_cphFormDetail_logmain_Password').send_keys(password)
    driver.find_element(By.ID, 'cphContent_cphFormDetail_logmain_LoginButton').click()

while True:
    sg.theme(theme_name_list[random.randint(0, len(theme_name_list))])
    #define layout
    layout=[[sg.Frame(' Select County ',[[sg.Radio('Montgomery', default=True, key="mon",group_id='2', font = 'Ubuntu')],[sg.Radio('Rutherford', default=False, key="rut",group_id='2', font = 'Ubuntu')],[sg.Radio('Robertson', default=False, key="rob",group_id='2', font = 'Ubuntu')],[sg.Radio('Sumner', default=False, key="sum",group_id='2', font = 'Ubuntu')],[sg.Radio('Williamson', default=False, key="will",group_id='2', font = 'Ubuntu')],[sg.Radio('Wilson', default=False, key="wil",group_id='2', font = 'Ubuntu')]],border_width=3,font = 'Ubuntu',relief = "solid")],
            [[sg.Frame(' Select Mode ',[[sg.Radio('General Session', default=True, key='gen',group_id=1,font = 'Ubuntu')],[sg.Radio('Circuit Courts', default=True, key='cir',group_id=1,font = 'Ubuntu')]],border_width=3,font = 'Ubuntu',relief = "solid")]],
            [[sg.Frame(' Date Option ',[[sg.Text('Enter the date range from',size=(20, 1), font='Ubuntu',justification='left')],[sg.Input(key='from', size=(20,1)), sg.CalendarButton('Calendar1',font="Ubuntu",  target='from', default_date_m_d_y=(int(today[1]),int(today[2]),int(today[0])), )],[sg.Text('Enter the date range to',size=(20, 1), font='Ubuntu',justification='left')],[sg.Input(key='to', size=(20,1)), sg.CalendarButton('Calendar2',font="Ubuntu",  target='to', default_date_m_d_y=(int(today[1]),int(today[2]),int(today[0])), )]],border_width=3,font = 'Ubuntu',relief = "solid")]],
            [[sg.Frame(' Login Credentials ',[[sg.Text('Username', font='Ubuntu',justification='left')],[sg.Input(key='user',font= 'Ubuntu')],[sg.Text('Password', font='Ubuntu',justification='left')],[sg.Input(key='pass',font= 'Ubuntu')]],border_width=3,key='newopt',font = 'Ubuntu',relief = "solid",visible=True)]],
            [sg.Button('START', font=('Ubuntu',12)),sg.Button('CANCEL', font=('Ubuntu',12))]]
    #Define Window
    win =sg.Window('TNcrtinfo',layout)
    #Read  values entered by user
    e,v=win.read()
    con = False 
    print(e,v)
    if e == None or e == "CANCEL":
        print('ham')
        print("exit")
        win.close()
        con = True
        print(1)
        break
    else:
        if  v['to'] == None or v['to'] == '' or v['from'] == None or v['from'] == '':
            print('Enter the date correctly')
            
            win.close()
            continue
        elif v['mon'] == False and v['rut'] == False and v['rob'] == False and v['sum'] == False and v['will'] == False and v['wil'] == False:
            print('please select the radio button')
            win.close()
            continue
        else:
            win.close()
            break
if con:
    pass
else:
    USER_INP_FROM = f"{v['from'].split(' ')[0].split('-')[1]}/{v['from'].split(' ')[0].split('-')[2]}/{v['from'].split(' ')[0].split('-')[0]}"
    USER_INP_TO = f"{v['to'].split(' ')[0].split('-')[1]}/{v['to'].split(' ')[0].split('-')[2]}/{v['to'].split(' ')[0].split('-')[0]}"
    dir_path = os.path.dirname(os.path.realpath('__file__'))
    
    if '/' in USER_INP_TO and '/' in USER_INP_FROM:
        USER_INP_YEAR_FROM = USER_INP_FROM.split('/')[-1]
        USER_INP_YEAR_TO = USER_INP_TO.split('/')[-1]
        
    current_court = []
    for key in v:
        if key == 'mon' or key=='rut' or key == 'rob' or key == 'sum' or key =='will' or key=='wil':
            if v[key] == True:
                current_county=key
                current_court.append(court_names[key])
        if key == 'gen' or key== 'cir':
            if v[key] == True:
                current_mode=key
                
    county=court_names[current_county]
    mode=mode_names[current_mode]
    username=v['user']
    password=v['pass']
    print(USER_INP_YEAR_FROM)
    print( USER_INP_YEAR_TO)
    print(county)
    print(mode)
    print(username)
    print(password)
    if (username=='') or (password==''):
        login=False
    else:
        login=True
    print(login)
    
    ##############################################################################################################
    #DATA EXTRACTION

    for i in current_court:
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get(f"https://{i}.tncrtinfo.com/cvCaseList.aspx?search=number")
        driver.implicitly_wait(30)
        today = date.today()
        x = random.randint(1, 1000000)
            
        if login == True:
            courts_login()

        driver.find_element(By.ID, 'ddlCourt').send_keys(mode)
        driver.find_element(By.LINK_TEXT, 'Civil').click()
        driver.find_element(By.LINK_TEXT, 'Search by Case Number').click()

        links = []
        print('Gathering Links...')

        for j in range(int(USER_INP_YEAR_FROM), int(USER_INP_YEAR_TO)+1):
            driver.get(f"https://{i}.tncrtinfo.com/cvCaseList.aspx?search=number")
            driver.find_element(by=By.ID, value="cphContent_cphSelectionCriteria_txtCaseYear").send_keys(j)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "cphContent_cphSelectionCriteria_cmdFindNow"))).click()

            while True:

                links += [i.find_element(By.XPATH, "td[2]/a").get_attribute('href') for i in driver.find_elements(By.XPATH, "//table[@class='searchList']/tbody/tr") if ('Plaintiff' in i.text and datetime.strptime(USER_INP_FROM, '%m/%d/%Y')  <= datetime.strptime(i.text.split(' ')[-1], '%m/%d/%Y') <= datetime.strptime(USER_INP_TO, '%m/%d/%Y'))]

                try:
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "cphContent_cphContentPaging_nextpage"))).click()
                except:
                    break

        driver.close()
        list1, list2, list3, list4 = links[:len(links)//4], links[len(links)//4:len(links)//4*2], links[len(links)//4*2:len(links)//4*3], links[len(links)//4*3:]
        #########################################################################################################
        
        others, divorce_detainer = [], []
        print('Extracting Plaintiffs...')
        for j in [list1, list2, list3, list4]:
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.get(f"https://{i}.tncrtinfo.com/cvCaseList.aspx?search=number")
            
            if login == True:
                courts_login()

            driver.find_element(By.ID, 'ddlCourt').send_keys(mode)
            driver.find_element(By.LINK_TEXT, 'Civil').click()
            
            for link in j:
                driver.get(link)
                try:
                    name = driver.find_element(By.ID, 'cphContent_cphFormDetail_frmdetail_spnparty').text
                except:
                    try:
                        if (driver.find_element(By.XPATH, "//*[contains(text(), 'Whoops!  Looks like the selected party does not exist')]")):
                            continue
                    except:
                        pass
                    name = '-'
                try:
                    case_style = driver.find_element(By.CLASS_NAME, 'titlegen').text
                except:
                    case_style_style = '-'
                try:
                    filing_date = driver.find_element(By.CLASS_NAME, 'field').text
                except:
                    filing_date = '-'
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Party Info'))).click()
                    address = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CLASS_NAME, 'addressinfo'))).text.replace('\n', ' ')
                except:
                    address = '-'
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Filings'))).click()
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Filing For')]")))
                    filing_type_list = [i.text for i in driver.find_elements(By.XPATH, "//table[@class='searchList']/tbody/tr/td[1]")]
                    filing_type = " || ".join(filing_type_list)
                except:
                    filing_type = "-"
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Rule Docket'))).click()
                    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Entry')]")))
                    details_list = [i.text.replace('\n', ' ') for i in driver.find_elements(By.XPATH, "//table[@class='searchList']/tbody/tr/td[3]")]
                    details = " || ".join(details_list)
                except:
                    details = "-" 
                print(f"NAME: {name}, CASE STYLE : {case_style}, DATE: {filing_date}, ADDRESS: {address}, TYPE: {filing_type}, DETAILS: {details}")
                print('--------------------------------------------------')
                print()
                
                if "Divorce" in filing_type or 'Detainer Warrant' in filing_type:
                    divorce_detainer.append([name, case_style, filing_date, filing_type, address, details])
                else:
                    others.append([name, case_style, filing_date, filing_type, address, details])

            driver.close()
        #########################################################################################################
        # EXCEL WRITE
        
        print('Generating EXCEL file...')
        def template():
            bold_font = Font(bold=True)
            center_aligned_text = Alignment(horizontal="center")
            double_border_side = Side(border_style="double")
            square_border = Border(top=double_border_side, right=double_border_side, bottom=double_border_side, left=double_border_side)

            sheet["A1"] = "NAME"
            sheet["A1"].font = bold_font
            sheet["A1"].alignment = center_aligned_text
            sheet["A1"].border = square_border
            sheet["B1"] = "CASE STYLE"
            sheet["B1"].font = bold_font
            sheet["B1"].alignment = center_aligned_text
            sheet["B1"].border = square_border
            sheet["C1"] = "FILLING DATE"
            sheet["C1"].font = bold_font
            sheet["C1"].alignment = center_aligned_text
            sheet["C1"].border = square_border
            sheet["D1"] = "FILLING TYPE"
            sheet["D1"].font = bold_font
            sheet["D1"].alignment = center_aligned_text
            sheet["D1"].border = square_border
            sheet["E1"] = "ADDRESS"
            sheet["E1"].font = bold_font
            sheet["E1"].alignment = center_aligned_text
            sheet["E1"].border = square_border
            sheet["F1"] = "DETAILS"
            sheet["F1"].font = bold_font
            sheet["F1"].alignment = center_aligned_text
            sheet["F1"].border = square_border

            dim_holder = DimensionHolder(worksheet=sheet)
            for col in range(sheet.min_column, sheet.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)
            sheet.column_dimensions = dim_holder

        workbook = Workbook()

        sheet = workbook.active
        template()
        for p,q in enumerate(others):
            sheet[f"A{p+2}"] = q[0]
            sheet[f"A{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"B{p+2}"] = q[1]
            sheet[f"B{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"C{p+2}"] = q[2]
            sheet[f"C{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"D{p+2}"] = q[3]
            sheet[f"D{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"E{p+2}"] = q[4]
            sheet[f"E{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"F{p+2}"] = q[5]
            sheet[f"F{p+2}"].alignment = Alignment(horizontal="center")

        sheet = workbook.create_sheet('Detainer or Divorce')
        template()
        for p,q in enumerate(divorce_detainer):
            sheet[f"A{p+2}"] = q[0]
            sheet[f"A{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"B{p+2}"] = q[1]
            sheet[f"B{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"C{p+2}"] = q[2]
            sheet[f"C{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"D{p+2}"] = q[3]
            sheet[f"D{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"E{p+2}"] = q[4]
            sheet[f"E{p+2}"].alignment = Alignment(horizontal="center")
            sheet[f"F{p+2}"] = q[5]
            sheet[f"F{p+2}"].alignment = Alignment(horizontal="center")

        workbook.save(f'{i}{today}{x}.xlsx')
        print(f"Excel file generated named {i}{today}{x}.xlsx")


# In[ ]:




